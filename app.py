import streamlit as st
import anthropic
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import docx
import io
import re
from datetime import datetime

# הגדרות עמוד
st.set_page_config(
    page_title="K2P - מערכת בדיקת מטלות",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS מעוצב וצבעוני
st.markdown("""
<style>
    /* ניקוי */
    #MainMenu, footer, header {visibility: hidden;}
    .block-container {padding-top: 1rem;}
    
    /* רקע */
    .main {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    }
    
    /* כותרת צבעונית */
    .custom-header {
        background: linear-gradient(90deg, #0080C8 0%, #7FBA00 50%, #0080C8 100%);
        padding: 2rem 3rem;
        border-radius: 15px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        margin-bottom: 2rem;
        text-align: center;
    }
    
    .custom-header h1 {
        color: white;
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
    }
    
    .custom-header h2 {
        color: rgba(255,255,255,0.95);
        font-size: 1.3rem;
        font-weight: 400;
        margin: 0.5rem 0 0 0;
    }
    
    /* ריבוע העלאה במרכז */
    .upload-container {
        max-width: 600px;
        margin: 0 auto;
        padding: 2rem;
    }
    
    [data-testid="stFileUploader"] {
        border: 3px dashed #0080C8;
        border-radius: 20px;
        background: white;
        padding: 3rem 2rem;
        box-shadow: 0 8px 20px rgba(0,0,0,0.1);
        transition: all 0.3s ease;
    }
    
    [data-testid="stFileUploader"]:hover {
        border-color: #7FBA00;
        transform: translateY(-5px);
        box-shadow: 0 12px 30px rgba(0,128,200,0.2);
    }
    
    [data-testid="stFileUploader"] section {
        padding: 2rem;
        text-align: center;
    }
    
    [data-testid="stFileUploader"] label {
        font-size: 1.2rem !important;
        font-weight: 600 !important;
        color: #0080C8 !important;
    }
    
    /* כפתור צבעוני */
    .stButton>button {
        background: linear-gradient(90deg, #0080C8 0%, #7FBA00 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.8rem 2.5rem;
        font-size: 1.1rem;
        font-weight: 600;
        box-shadow: 0 4px 15px rgba(0,128,200,0.3);
        transition: all 0.3s ease;
        width: 100%;
        max-width: 300px;
        margin: 0 auto;
        display: block;
    }
    
    .stButton>button:hover {
        transform: translateY(-3px);
        box-shadow: 0 6px 20px rgba(0,128,200,0.4);
    }
    
    /* הגדרות */
    .streamlit-expanderHeader {
        background: white;
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        font-weight: 600;
        padding: 1rem;
    }
    
    /* הודעות צבעוניות */
    .stSuccess {
        background: linear-gradient(90deg, #d4edda 0%, #c3e6cb 100%);
        border-left: 4px solid #28a745;
        border-radius: 10px;
        padding: 1rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    
    .stInfo {
        background: linear-gradient(90deg, #d1ecf1 0%, #bee5eb 100%);
        border-left: 4px solid #17a2b8;
        border-radius: 10px;
        padding: 1rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    
    .stError {
        background: linear-gradient(90deg, #f8d7da 0%, #f5c6cb 100%);
        border-left: 4px solid #dc3545;
        border-radius: 10px;
        padding: 1rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    
    /* מטריקות */
    [data-testid="stMetricValue"] {
        font-size: 2rem;
        font-weight: 700;
        background: linear-gradient(90deg, #0080C8 0%, #7FBA00 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    
    [data-testid="stMetricLabel"] {
        font-size: 0.9rem;
        color: #666;
        font-weight: 600;
    }
    
    /* קלפי תוצאות */
    .results-card {
        background: white;
        border-radius: 15px;
        padding: 2rem;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        margin: 1rem 0;
    }
    
    /* טבלה */
    table {
        background: white;
        border-radius: 10px;
        overflow: hidden;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    
    table thead {
        background: linear-gradient(90deg, #0080C8 0%, #7FBA00 100%);
    }
    
    table th {
        color: white !important;
        font-weight: 600;
        padding: 15px !important;
    }
    
    table td {
        padding: 12px !important;
        border-bottom: 1px solid #f0f0f0;
    }
    
    /* תיבת טקסט */
    .stTextInput>div>div>input {
        border-radius: 8px;
        border: 2px solid #e0e0e0;
        padding: 0.7rem;
    }
    
    .stTextInput>div>div>input:focus {
        border-color: #0080C8;
        box-shadow: 0 0 0 3px rgba(0,128,200,0.1);
    }
    
    /* Progress bar */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #0080C8 0%, #7FBA00 100%);
    }
    
    /* לוגו */
    .logo-container {
        text-align: center;
        margin-bottom: 2rem;
    }
    
    .logo-container img {
        filter: drop-shadow(0 4px 8px rgba(0,0,0,0.1));
    }
</style>
""", unsafe_allow_html=True)

# לוגו במרכז
st.markdown('<div class="logo-container">', unsafe_allow_html=True)
try:
    st.image("k2p_logo.png", width=200)
except:
    st.markdown("<h3 style='text-align:center;color:#0080C8;'>K2P</h3>", unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# כותרת צבעונית
st.markdown("""
<div class="custom-header">
    <h1>📚 מערכת בדיקת מטלות אקדמאיות</h1>
    <h2>🎓 קורס התנהגות ארגונית</h2>
</div>
""", unsafe_allow_html=True)

# API Key
if 'api_key' not in st.session_state:
    st.session_state.api_key = ""

# הגדרות
with st.expander("⚙️ הגדרות", expanded=False):
    api_key = st.text_input(
        "Claude API Key",
        type="password",
        value=st.session_state.api_key,
        placeholder="הזן API Key",
        key="api_input"
    )
    if api_key:
        st.session_state.api_key = api_key
        st.success("✅ API Key נשמר בהצלחה!")

st.markdown("<br>", unsafe_allow_html=True)

# פונקציות
def read_docx(file):
    try:
        doc = docx.Document(file)
        return '\n'.join([p.text for p in doc.paragraphs if p.text.strip()])
    except Exception as e:
        return f"שגיאה: {str(e)}"

def extract_work_number(filename):
    """חילוץ מספר מטלה בלבד מהשם הקובץ"""
    name = filename.replace('.docx', '').replace('.doc', '')
    
    match = re.search(r'WorkCode[_-]?(\d+)', name, re.IGNORECASE)
    if match:
        return match.group(1)
    
    match = re.search(r'\b(\d{8,9})\b', name)
    if match:
        return match.group(1)
    
    match = re.search(r'\b(\d{4,})\b', name)
    if match:
        return match.group(1)
    
    match = re.search(r'(\d+)', name)
    if match:
        return match.group(1)
    
    return ""

def grade_assignment(content, filename, api_key):
    try:
        client = anthropic.Anthropic(api_key=api_key)
        
        prompt = f"""אתה בודק מטלות בקורס התנהגות ארגונית. בדוק לפי המחוון:

**מחוון (100 נק'):**

שאלה 1 - תרבות (40):
- א (15): תרבות כללית = המדינה. אם חסר לגמרי → 15-
- ב (15): תרבות ארגונית. אם חסר פירוט → 5-
- ג (10): יחסי גומלין. אם חסר → 10-

שאלה 2 - מבנה (20): 3 תיאוריות
שאלה 3 - תהליך (20): 2 תיאוריות
שאלה 4 - תוכן (20): 2 תיאוריות

**הפחתת נקודות:**
- "ניתן להרחיב" או "חסר פירוט קל" → 2-3 נקודות
- חסר דבר משמעותי → 5-15 נקודות

**חשוב מאוד - כתיבת הערות:**
1. כתוב **רק** מה שחסר או חלש
2. אל תכתוב "עבודה טובה", "מצוין", "כל הכבוד", "נעשה יפה" - שום דבר חיובי!
3. אל תכתוב "הסטודנט", "כתב", "לא הבין"
4. כל הערה בשורה נפרדת
5. **חובה**: כתוב את ההפחתה בסוגריים בסוף כל הערה
6. אם אין מה לכתוב - השאר ריק (אל תכתוב כלום!)
7. תהיה נדיב - רוב הציונים 85-95

**פורמט נכון:**
"שאלה 1: חסרה תרבות כללית - תרבות המדינה (-15)
שאלה 3: ניתן להרחיב על מוטיבציה (-2)"

**פורמט לא נכון (אסור!):**
"עבודה טובה מאוד! רק..."
"הסטודנט כתב יפה אבל..."
"מצוין! חסר רק..."

JSON:
{{
  "workNumber": "מספר",
  "grade": 0-100,
  "comments": "הערות או ריק"
}}

קובץ: {filename}
תוכן: {content[:12000]}"""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4000,
            messages=[{"role": "user", "content": prompt}]
        )
        
        response_text = message.content[0].text
        json_match = re.search(r'\{[\s\S]*\}', response_text)
        
        if json_match:
            import json
            result = json.loads(json_match.group(0))
            return result
        
        return {
            "workNumber": extract_work_number(filename),
            "grade": 0,
            "comments": "לא הצלחתי לפענח תשובה"
        }
        
    except Exception as e:
        return {
            "workNumber": extract_work_number(filename),
            "grade": 0,
            "comments": f"שגיאה: {str(e)}"
        }

def create_styled_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "תוצאות בדיקה"
    
    headers = ['שם קובץ', 'מספר', 'ציון', 'הערות']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="0080C8", end_color="7FBA00", fill_type="solid")
    header_font = Font(bold=True, size=12, name="Arial", color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, 5):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    row_colors = ["E6F2FF", "E8F5E9", "FFF9E6", "F3E5F5", "FFE6F0", "E1F5FE"]
    
    def get_grade_color(grade):
        if grade >= 90: return "C8E6C9"
        if grade >= 85: return "BBDEFB"
        if grade >= 80: return "FFF59D"
        if grade >= 70: return "FFCC80"
        return "FFCDD2"
    
    for idx, result in enumerate(results):
        row_num = idx + 2
        bg_color = row_colors[idx % len(row_colors)]
        
        ws.append([
            result['filename'],
            result['workNumber'],
            result['grade'],
            result['comments']
        ])
        
        for col in range(1, 5):
            cell = ws.cell(row_num, col)
            
            if col == 3:
                cell.fill = PatternFill(start_color=get_grade_color(result['grade']), 
                                       end_color=get_grade_color(result['grade']), 
                                       fill_type="solid")
                cell.font = Font(bold=True, size=16, name="Arial")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                cell.font = Font(size=11, name="Arial")
                
                if col == 2:
                    cell.font = Font(bold=True, size=12, name="Arial")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            
            cell.border = Border(
                left=Side(style='thin', color="CCCCCC"),
                right=Side(style='thin', color="CCCCCC"),
                top=Side(style='thin', color="CCCCCC"),
                bottom=Side(style='thin', color="CCCCCC")
            )
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 100
    
    ws.row_dimensions[1].height = 30
    for idx, result in enumerate(results):
        row_num = idx + 2
        lines = len(result['comments'].split('\n')) if result['comments'] else 1
        ws.row_dimensions[row_num].height = max(60, lines * 20 + 10)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ריבוע העלאה במרכז
st.markdown('<div class="upload-container">', unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "📤 גרור קבצים לכאן או לחץ לבחירה",
    type=['docx'],
    accept_multiple_files=True,
    help="תומך ב-Word (.docx) בלבד"
)

st.markdown('</div>', unsafe_allow_html=True)

if uploaded_files:
    st.success(f"✅ {len(uploaded_files)} קבצים הועלו בהצלחה!")
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([2, 1, 2])
    with col2:
        if st.button("🚀 התחל בדיקה", type="primary"):
            if not st.session_state.api_key:
                st.error("❌ נא להזין Claude API Key בהגדרות")
            else:
                results = []
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                for idx, file in enumerate(uploaded_files):
                    status_text.text(f"בודק מטלה {idx + 1} מתוך {len(uploaded_files)}...")
                    progress_bar.progress((idx + 1) / len(uploaded_files))
                    
                    content = read_docx(file)
                    result = grade_assignment(content, file.name, st.session_state.api_key)
                    results.append({
                        'filename': file.name,
                        'workNumber': result.get('workNumber', ''),
                        'grade': result.get('grade', 0),
                        'comments': result.get('comments', '')
                    })
                
                st.session_state.results = results
                progress_bar.empty()
                status_text.empty()
                st.success("✅ הבדיקה הושלמה בהצלחה!")
                st.rerun()

# תוצאות
if 'results' in st.session_state and st.session_state.results:
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown('<div class="results-card">', unsafe_allow_html=True)
    st.markdown("### 📊 תוצאות הבדיקה")
    
    grades = [r['grade'] for r in st.session_state.results]
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("ממוצע", f"{sum(grades)/len(grades):.1f}")
    with col2:
        st.metric("מקסימום", f"{max(grades)}")
    with col3:
        st.metric("מינימום", f"{min(grades)}")
    with col4:
        st.metric("סה״כ", f"{len(grades)}")
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # טבלה
    table_html = "<table style='width:100%;'>"
    table_html += "<thead><tr>"
    table_html += "<th style='text-align: right;'>קובץ</th>"
    table_html += "<th style='text-align: center;'>מספר</th>"
    table_html += "<th style='text-align: center;'>ציון</th>"
    table_html += "<th style='text-align: right;'>הערות</th>"
    table_html += "</tr></thead><tbody>"
    
    for idx, r in enumerate(st.session_state.results):
        bg = "#fafafa" if idx % 2 == 0 else "#ffffff"
        
        if r['grade'] >= 90:
            grade_color = "#c8e6c9"
        elif r['grade'] >= 85:
            grade_color = "#bbdefb"
        elif r['grade'] >= 80:
            grade_color = "#fff59d"
        else:
            grade_color = "#ffcdd2"
            
        table_html += f"<tr style='background-color: {bg};'>"
        table_html += f"<td style='text-align: right;'>{r['filename']}</td>"
        table_html += f"<td style='text-align: center; font-weight: 600;'>{r['workNumber']}</td>"
        table_html += f"<td style='text-align: center; background-color: {grade_color}; font-weight: 700; font-size: 1.2rem;'>{r['grade']}</td>"
        table_html += f"<td style='text-align: right; white-space: pre-line; font-size: 0.9rem; color: #555;'>{r['comments']}</td>"
        table_html += "</tr>"
    
    table_html += "</tbody></table>"
    st.markdown(table_html, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([2, 1, 2])
    
    with col1:
        excel_file = create_styled_excel(st.session_state.results)
        st.download_button(
            label="📥 הורד קובץ Excel",
            data=excel_file,
            file_name=f"תוצאות_K2P_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    with col3:
        if st.button("🗑️ נקה תוצאות", use_container_width=True):
            del st.session_state.results
            st.rerun()

st.markdown("<br><br>", unsafe_allow_html=True)
st.markdown("<div style='text-align:center;color:#888;font-size:0.9rem;padding:2rem;'>K2P - Knowledge to People • גרסה 2.0 • Powered by Claude AI</div>", unsafe_allow_html=True)
